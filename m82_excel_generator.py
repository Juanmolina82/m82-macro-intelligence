# -*- coding: utf-8 -*-
import os
import sys

# Asegurar instalación automática de openpyxl
try:
    import openpyxl
except ImportError:
    print("[INFO] openpyxl no detectado. Instalando dependencia...")
    os.system("pip install openpyxl")
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_reporte_excel():
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # RESUMEN EJECUTIVO
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.views.sheetView[0].showGridLines = True
    
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    fill_accent = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True, color="1A365D")
    font_regular = Font(name="Arial", size=10, color="2D3748")
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')
    )
    
    ws1["B2"] = "M82 WORLD SPY SYSTEM - MATRIZ DE CONTROL FINANCIERO"
    ws1["B2"].font = Font(name="Arial", size=14, bold=True, color="1A365D")
    ws1["B3"] = "Infraestructura Cloud Sincronizada // Cierre Intermercado"
    ws1["B3"].font = Font(name="Arial", size=10, italic=True, color="4A5568")
    
    ws1["B5"] = "Fideicomiso de Custodia Privada (Molina Holdings LLC):"
    ws1["B5"].font = font_regular
    ws1["F5"] = 2330000000
    ws1["F5"].font = font_bold
    ws1["F5"].number_format = "$#,##0"
    
    ws1["B6"] = "Flujo Neto Consolidado XLK (Tecnologia):"
    ws1["B6"].font = font_regular
    ws1["F6"] = 505000000
    ws1["F6"].font = font_bold
    ws1["F6"].number_format = "$#,##0"
    
    # ---------------------------------------------------------
    # TAXONOMÍA LSEG WORKSPACE
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="Taxonomia LSEG")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = ["Categoria de Industria M82", "Metrica de Amplitud", "Parametro Tecnico", "Estado"]
    for col_num, header in enumerate(headers, 2):
        cell = ws2.cell(row=3, column=col_num, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin
        
    sub_sectores = [
        "Communication Equipment", "Computer Hardware", "Consumer Electronics",
        "Electronic Components", "Electronics & Computer Distribution",
        "Information Technology Services", "Scientific & Technical Instruments",
        "Semiconductor Equipment & Materials", "Semiconductors",
        "Software - Application", "Software - Infrastructure", "Solar"
    ]
    
    for idx, sector in enumerate(sub_sectores, 4):
        c_sec = ws2.cell(row=idx, column=2, value=sector)
        c_amp = ws2.cell(row=idx, column=3, value=16.89 if idx == 5 else 0.00)
        c_param = ws2.cell(row=idx, column=4, value="LEADING" if idx in [5, 12, 13] else "Alineacion")
        c_state = ws2.cell(row=idx, column=5, value="Estable")
        
        for cell in [c_sec, c_amp, c_param, c_state]:
            cell.font = font_regular
            cell.border = border_thin
            if idx % 2 == 0:
                cell.fill = fill_zebra
            if idx == 5: # Rally de Dell Technologies
                cell.fill = fill_accent
                cell.font = Font(name="Arial", size=10, bold=True, color="2B6CB0")
                
        c_amp.number_format = "0.00"
        c_amp.alignment = Alignment(horizontal="right")
        c_param.alignment = Alignment(horizontal="center")
        c_state.alignment = Alignment(horizontal="center")
        
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    filename = "M82_Sector_Rotation_Matrix.xlsx"
    wb.save(filename)
    print(f"\n[SUCCESS] Archivo Excel compilado nativamente: {filename}")

if __name__ == "__main__":
    generar_reporte_excel()
