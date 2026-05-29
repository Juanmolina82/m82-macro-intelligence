# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def actualizar_excel_snowflake():
    excel_file = "M82_Sector_Rotation_Matrix.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"[ERROR] No se encontró la matriz base {excel_file}. Ejecuta primero tu generador maestro.")
        return

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Taxonomia LSEG"]
    
    # Colores y Estilos para el nuevo catalizador
    fill_snowflake = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
    font_highlight = Font(name="Arial", size=10, bold=True, color="2B6CB0")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')
    )

    # Buscar e inyectar datos en las filas correspondientes a Software de LSEG Workspace
    # Fila 13: Software - Application // Fila 14: Software - Infrastructure
    row_targets = [13, 14]
    
    for row in row_targets:
        if ws.cell(row=row, column=2).value in ["Software - Application", "Software - Infrastructure"]:
            ws.cell(row=row, column=3, value=34.00) # Crecimiento récord Snowflake
            ws.cell(row=row, column=4, value="LEADING")
            ws.cell(row=row, column=5, value="Rally SNOW +40% / Contrato AWS $6B")
            
            # Aplicar estilos estilizados
            for col in range(2, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_snowflake
                cell.font = font_highlight
                cell.border = border_thin
                if col == 3:
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col in [4, 5]:
                    cell.alignment = Alignment(horizontal="center")

    wb.save(excel_file)
    print(f"[SUCCESS] Matriz de sectores en Excel actualizada de forma inmutable: {excel_file}")

if __name__ == "__main__":
    actualizar_excel_snowflake()
